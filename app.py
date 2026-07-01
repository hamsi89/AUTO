import streamlit as st
import pandas as pd
import datetime
import os
import io
import shutil
import openpyxl  
import json
import gspread

# [필수] 페이지 설정은 앱 전체 최상단에 딱 1번만 실행되어야 오류가 발생하지 않습니다.
st.set_page_config(page_title="VINI COFFEE 재고관리 시스템", page_icon="☕", layout="wide")

# 1. Secrets에서 문자열 안전하게 가져오기
raw_creds = st.secrets.get("gcp_credentials_json", "")
raw_auth = st.secrets.get("gcp_authorized_user_json", "")

# 디버깅용 안전장치: 데이터가 비어있다면 에러 메시지 출력 후 즉시 중단
if not raw_creds or not raw_auth:
    st.error("Streamlit Secrets에 구글 인증 정보가 등록되지 않았거나 비어있습니다.")
    st.stop()

try:
    # .strip()을 붙여서 앞뒤의 무의미한 줄바꿈이나 공백을 완전히 제거합니다.
    creds_dict = json.loads(raw_creds.strip())
    auth_user_dict = json.loads(raw_auth.strip())
except json.decoder.JSONDecodeError as e:
    st.error("⚠️ 구글 인증서 JSON 형식이 올바르지 않습니다. Secrets 입력을 확인해 주세요.")
    st.warning(f"읽어온 데이터 앞부분: {raw_creds[:30]}...")
    st.stop()

# 2. gspread 인증 진행
gc = gspread.oauth_from_dict(creds_dict, auth_user_dict)

# 1. 세션 상태(Session State)에 로그인 유무 저장할 공간 초기화
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

# 2. 로그인하지 않은 상태일 때 -> 로그인 화면 표시 후 백엔드 전면 차단
if not st.session_state.logged_in:
    st.title("🔒 매장 재고 관리 시스템")
    st.subheader("관리자 인증이 필요합니다.")
    
    password_input = st.text_input("관리자 비밀번호를 입력하세요", type="password")
    login_button = st.button("로그인")

    if login_button or password_input:
        if password_input == "122000":
            st.session_state.logged_in = True
            st.success("인증에 성공했습니다! 잠시만 기다려주세요...")
            st.rerun()  
        else:
            st.error("비밀번호가 올바르지 않습니다. 다시 입력해주세요.")
    st.stop()  # 로그인 성공 전까지 하단 코드 로드 방지

# =========================================================================
# 로그인 성공 상태일 때만 정상 실행되는 구역 (재고 관리 핵심 로직)
# =========================================================================

# 파일 이름 설정
ORIGINAL_EXCEL_PATH = "VINI_COFFEE_통합_식자재_및_매출관리_시스템_v3_주간체크리스트추가.xlsx"
CUSTOM_MASTER_FILE = "vini_custom_master.csv"  
STOCK_LOG_FILE = "vini_daily_stock_log.csv"

# 🍷 와인 전용 파일 및 로그 설정
WINE_EXCEL_PATH = "와인_입출고양식_디자인적용_현재고요약.xlsx"
WINE_MASTER_FILE = "wine_custom_master.csv"
WINE_LOG_FILE = "wine_daily_stock_log.csv"

# 원본 엑셀의 진짜 시트 이름 지정
SHEET_MAP = {
    "원재료": "원재료",
    "부자재": "부자재",
    "디저트&완제품": "디저트&완제품"
}

@st.cache_data
def load_excel_master():
    """기존 엑셀 시트에서 최초 마스터 품목 정보를 파싱하는 함수"""
    sheets_to_try = {
        "원재료": ["원재료", "원재료(간략)"],
        "부자재": ["부자재", "부자재(간략)"],
        "디저트&완제품": ["디저트&완제품", "디저트&완제품(간략)"]
    }
    
    master_df_list = []
    
    if os.path.exists(ORIGINAL_EXCEL_PATH):
        try:
            xl = pd.ExcelFile(ORIGINAL_EXCEL_PATH)
            for cat, standard_names in sheets_to_try.items():
                target_sheet = None
                for name in xl.sheet_names:
                    if name.strip() in standard_names or cat in name:
                        target_sheet = name
                        break
                
                if target_sheet:
                    df = pd.read_excel(xl, sheet_name=target_sheet, skiprows=2)
                    df.columns = [str(c).strip().replace(" ", "") for c in df.columns]
                    
                    name_col = None
                    for col in df.columns:
                        if '품목이름' in col or '구분' in col or '품목명' in col:
                            name_col = col
                            break
                    
                    if name_col:
                        df = df.dropna(subset=[name_col])
                        df = df[df[name_col].astype(str).str.strip() != '0']
                        df = df[df[name_col].astype(str).str.strip() != '']
                        
                        expiry_col = [c for c in df.columns if '유통' in c]
                        stock_col = [c for c in df.columns if '재고' in c]
                        
                        temp_df = pd.DataFrame()
                        temp_df['품목명'] = df[name_col].astype(str).str.strip()
                        temp_df['대분류'] = cat
                        
                        if expiry_col:
                            temp_df['유통기한'] = df[expiry_col[0]].astype(str).str.strip()
                        else:
                            temp_df['유통기한'] = ""
                            
                        if stock_col:
                            temp_df['엑셀기본재고'] = pd.to_numeric(df[stock_col[0]], errors='coerce').fillna(0).astype(int)
                        else:
                            temp_df['엑셀기본재고'] = 0
                            
                        master_df_list.append(temp_df)
                        
            if master_df_list:
                return pd.concat(master_df_list, ignore_index=True)
        except Exception as e:
            st.error(f"엑셀 품목 로드 실패: {e}")
            
    return pd.DataFrame([
        {"품목명": "커피(퍼플-마스터피스)", "대분류": "원재료", "유통기한": "2027-04-13", "엑셀기본재고": 54},
        {"품목명": "서울우유", "대분류": "원재료", "유통기한": "2026-06-25", "엑셀기본재고": 42}
    ])

def get_latest_master_and_logs():
    if not os.path.exists(CUSTOM_MASTER_FILE):
        base_master = load_excel_master()
        base_master.to_csv(CUSTOM_MASTER_FILE, index=False, encoding='utf-8-sig')
        
    if not os.path.exists(STOCK_LOG_FILE):
        df_empty = pd.DataFrame(columns=["날짜", "대분류", "품목명", "구분", "수량", "유통기한"])
        df_empty.to_csv(STOCK_LOG_FILE, index=False, encoding='utf-8-sig')

    master = pd.read_csv(CUSTOM_MASTER_FILE, encoding='utf-8-sig')
    master['엑셀기본재고'] = master['엑셀기본재고'].fillna(0).astype(int)
    master['유통기한'] = master['유통기한'].fillna("").astype(str)
    
    logs = pd.read_csv(STOCK_LOG_FILE, encoding='utf-8-sig')
    
    if not logs.empty and "유통기한" in logs.columns:
        expiry_logs = logs[logs["유통기한"].notna() & (logs["유통기한"].astype(str) != "") & (logs["유통기한"].astype(str) != "nan")]
        if not expiry_logs.empty:
            latest_expiries = expiry_logs.sort_values(by="날짜").groupby("품목명").last()["유통기한"].to_dict()
            master["유통기한"] = master["품목명"].map(latest_expiries).fillna(master["유통기한"])
            
    return master, logs

# 🍷 와인 전용 마스터 및 로그 로드 함수
@st.cache_data
def load_wine_master():
    """와인 엑셀 파일에서 마스터 품목 리스트를 추출하는 함수"""
    if os.path.exists(WINE_EXCEL_PATH):
        try:
            xl = pd.ExcelFile(WINE_EXCEL_PATH)
            target_sheet = xl.sheet_names[0]
            for name in xl.sheet_names:
                if "와인" in name or "재고" in name:
                    target_sheet = name
                    break
            
            df = pd.read_excel(xl, sheet_name=target_sheet, skiprows=2)
            df.columns = [str(c).strip().replace(" ", "") for c in df.columns]
            
            name_col = None
            for col in df.columns:
                if '품목' in col or '와인명' in col or '이름' in col or '구분' in col:
                    name_col = col
                    break
                    
            if name_col:
                df = df.dropna(subset=[name_col])
                df = df[df[name_col].astype(str).str.strip() != '0']
                
                stock_col = [c for c in df.columns if '재고' in c or '이월' in c]
                expiry_col = [c for c in df.columns if '유통' in c or '빈티지' in c]
                
                temp_df = pd.DataFrame()
                temp_df['품목명'] = df[name_col].astype(str).str.strip()
                temp_df['대분류'] = "와인"
                temp_df['유통기한'] = df[expiry_col[0]].astype(str).str.strip() if expiry_col else ""
                temp_df['엑셀기본재고'] = pd.to_numeric(df[stock_col[0]], errors='coerce').fillna(0).astype(int) if stock_col else 0
                return temp_df
        except Exception as e:
            st.error(f"와인 엑셀 품목 로드 실패: {e}")
            
    return pd.DataFrame([
        {"품목명": "레드와인_A", "대분류": "와인", "유통기한": "", "엑셀기본재고": 10},
        {"품목명": "화이트와인_B", "대분류": "와인", "유통기한": "", "엑셀기본재고": 5}
    ])

def get_wine_master_and_logs():
    if not os.path.exists(WINE_MASTER_FILE):
        wine_master = load_wine_master()
        wine_master.to_csv(WINE_MASTER_FILE, index=False, encoding='utf-8-sig')
        
    if not os.path.exists(WINE_LOG_FILE):
        df_empty = pd.DataFrame(columns=["날짜", "대분류", "품목명", "구분", "수량", "유통기한"])
        df_empty.to_csv(WINE_LOG_FILE, index=False, encoding='utf-8-sig')

    master = pd.read_csv(WINE_MASTER_FILE, encoding='utf-8-sig')
    master['엑셀기본재고'] = master['엑셀기본재고'].fillna(0).astype(int)
    master['유통기한'] = master['유통기한'].fillna("").astype(str)
    logs = pd.read_csv(WINE_LOG_FILE, encoding='utf-8-sig')
    return master, logs

master_data, log_df = get_latest_master_and_logs()
wine_master_data, wine_log_df = get_wine_master_and_logs()

# 세션 상태 초기화
if "success_msg" not in st.session_state: st.session_state.success_msg = ""
if "warning_msg" not in st.session_state: st.session_state.warning_msg = ""
if "in_qty" not in st.session_state: st.session_state.in_qty = 0
if "out_qty" not in st.session_state: st.session_state.out_qty = 0
if "bulk_download_ready" not in st.session_state: st.session_state.bulk_download_ready = False
if "bulk_excel_bytes" not in st.session_state: st.session_state.bulk_excel_bytes = None
if "bulk_filename" not in st.session_state: st.session_state.bulk_filename = ""
if "orig_excel_bytes" not in st.session_state: st.session_state.orig_excel_bytes = None

# 와인 전용 세션 상태 초기화
if "wine_download_ready" not in st.session_state: st.session_state.wine_download_ready = False
if "wine_excel_bytes" not in st.session_state: st.session_state.wine_excel_bytes = None
if "wine_filename" not in st.session_state: st.session_state.wine_filename = ""
if "wine_orig_bytes" not in st.session_state: st.session_state.wine_orig_bytes = None

# 유통기한 임박 계산
today = datetime.date.today()
imminent_items = []
for idx, row in master_data.iterrows():
    utg = str(row['유통기한']).strip()
    if utg and utg != '0' and utg != '1899-12-31' and utg != 'nan' and utg != '':
        try:
            expiry_date = pd.to_datetime(utg).date()
            days_left = (expiry_date - today).days
            if 0 <= days_left <= 30:
                imminent_items.append(f"• **{row['품목명']}** ({days_left}일 남음)")
        except:
            pass

# --- 개별 입출고 마스터 반영 함수 ---
def update_excel_vini_rule(cat, item_name, qty, target_date, new_expiry=None, action="inbound"):
    if not os.path.exists(ORIGINAL_EXCEL_PATH):
        st.error(f"서버에 원본 엑셀 파일을 찾을 수 없습니다: {ORIGINAL_EXCEL_PATH}")
        return False
    try:
        wb = openpyxl.load_workbook(ORIGINAL_EXCEL_PATH)
        target_sheet = SHEET_MAP.get(cat, cat)
        if target_sheet not in wb.sheetnames:
            return False
        ws = wb[target_sheet]
        header_row = 3 
        
        name_col_idx, expiry_col_idx, target_col_idx = None, None, None
        target_day_num = str(target_date.day)          
        target_day_label = f"{target_date.day}일"       
        
        for col in range(1, ws.max_column + 1):
            cell_raw = ws.cell(row=header_row, column=col).value
            if cell_raw is None: continue
            val = str(cell_raw).strip().replace(" ", "")
            
            if '품목이름' in val or '품목명' in val or '구분' in val: name_col_idx = col
            elif '유통' in val: expiry_col_idx = col
            
            if action == "inbound" and ('금월입고' in val or '입고' in val): target_col_idx = col
            elif action == "outbound" and (val == target_day_label or val == target_day_num or val == f"0{target_day_num}"): target_col_idx = col

        if not name_col_idx:
            st.error("엑셀 헤더 구조에서 '품목이름' 컬럼을 찾지 못했습니다.")
            return False
        if not target_col_idx:
            msg = "금월 입고" if action == "inbound" else f"{target_day_label}"
            st.error(f"❌ 엑셀 3열에서 [{msg}] 관련 열을 매칭하지 못했습니다.")
            return False

        item_found = False
        for row in range(header_row + 1, ws.max_row + 1):
            cell_item_name = str(ws.cell(row=row, column=name_col_idx).value).strip()
            if cell_item_name == item_name:
                target_cell = ws.cell(row=row, column=target_col_idx)
                
                try: current_val = int(target_cell.value) if target_cell.value is not None else 0
                except: current_val = 0

                target_cell.value = current_val + qty
                if new_expiry and expiry_col_idx: ws.cell(row=row, column=expiry_col_idx).value = new_expiry
                item_found = True
                break
                
        if not item_found:
            st.error(f"엑셀 파일에서 [{item_name}] 품목을 찾을 수 없습니다.")
            return False

        wb.save(ORIGINAL_EXCEL_PATH)
        wb.close()
        return True
    except Exception as e:
        st.error(f"서버 엑셀 장부 반영 실패: {e}")
        return False

# --- 콜백 함수 구역 ---
def save_inbound_callback():
    qty = st.session_state.in_qty
    if qty > 0:
        cat = st.session_state.in_cat
        item = st.session_state.in_item
        in_date_val = st.session_state.in_date
        expiry_str = st.session_state.in_utg.strftime("%Y-%m-%d")
        
        excel_success = update_excel_vini_rule(cat, item, qty, in_date_val, new_expiry=expiry_str, action="inbound")
        
        if excel_success:
            current_logs = pd.read_csv(STOCK_LOG_FILE, encoding='utf-8-sig')
            new_data = pd.DataFrame([{
                "날짜": in_date_val.strftime("%Y-%m-%d"),
                "대분류": cat,
                "품목명": item,
                "구분": "금월 입고",
                "수량": qty,
                "유통기한": expiry_str
            }])
            current_logs = pd.concat([current_logs, new_data], ignore_index=True)
            current_logs.to_csv(STOCK_LOG_FILE, index=False, encoding='utf-8-sig')
            st.session_state.success_msg = f"📥 [입고 완료] 품목: {item} | 수량: {qty}개 ➡️ 엑셀 [금월 입고] 열에 안전하게 누적되었습니다."
            st.session_state.in_qty = 0  
            st.cache_data.clear()
    else:
        st.warning("⚠️ 입고 수량을 1개 이상 입력하셔야 합니다.")

def save_outbound_callback():
    qty = st.session_state.out_qty
    if qty > 0:
        cat = st.session_state.out_cat
        item = st.session_state.out_item
        out_date_val = st.session_state.out_date
        
        excel_success = update_excel_vini_rule(cat, item, qty, out_date_val, action="outbound")
        
        if excel_success:
            current_logs = pd.read_csv(STOCK_LOG_FILE, encoding='utf-8-sig')
            new_data = pd.DataFrame([{
                "날짜": out_date_val.strftime("%Y-%m-%d"),
                "대분류": cat,
                "품목명": item,
                "구분": "월 소모(출고)",
                "수량": qty,
                "유통기한": ""
            }])
            current_logs = pd.concat([current_logs, new_data], ignore_index=True)
            current_logs.to_csv(STOCK_LOG_FILE, index=False, encoding='utf-8-sig')
            st.session_state.success_msg = f"📤 [소모 완료] 품목: {item} | 소모: {qty}개 ➡️ 엑셀 [{out_date_val.day}일] 열에 정확히 합산되었습니다."
            st.session_state.out_qty = 0  
            st.cache_data.clear()
    else:
        st.warning("⚠️ 출고 수량을 1개 이상 입력하셔야 합니다.")


# --- 사이드바 메뉴 레이아웃 ---
st.sidebar.title("관리자 메뉴")
if st.sidebar.button("로그아웃"):
    st.session_state.logged_in = False
    st.rerun()

if imminent_items:
    with st.sidebar.expander("🚨 유통기한 임박 품목 알림", expanded=True):
        for item in imminent_items:
            st.warning(item)

menu = st.sidebar.radio(
    "메뉴 이동", 
    [
        "📥 물품 입고 등록 (개별)", 
        "📤 물품 출고 등록 (소모-개별)", 
        "📝 전품목 일괄 입력 (엑셀 스타일)",
        "🍷 와인 재고 관리 (일괄 입력)",
        "📋 실시간 현재고 현황판", 
        "📊 월별 수불 대장 및 백업 다운로드",
        "⚙️ 품목 추가/삭제 관리",
        "🛠️ 데이터 관리 및 엑셀 동기화"
    ]
)

# 메뉴 이동 시 상대 메뉴의 임시 버퍼 데이터를 강제로 비워 교차 오염을 원천 차단합니다.
if menu != "📝 전품목 일괄 입력 (엑셀 스타일)":
    st.session_state.bulk_download_ready = False
    st.session_state.bulk_excel_bytes = None
    st.session_state.orig_excel_bytes = None
if menu != "🍷 와인 재고 관리 (일괄 입력)":
    st.session_state.wine_download_ready = False
    st.session_state.wine_excel_bytes = None
    st.session_state.wine_orig_bytes = None

if st.session_state.success_msg:
    st.success(st.session_state.success_msg)
    st.session_state.success_msg = ""
if st.session_state.warning_msg:
    st.warning(st.session_state.warning_msg)
    st.session_state.warning_msg = ""

def show_today_logs_and_management(is_wine=False):
    st.markdown("---")
    st.subheader("🔍 오늘 실시간 입력된 내역")
    target_log_file = WINE_LOG_FILE if is_wine else STOCK_LOG_FILE
    current_logs = pd.read_csv(target_log_file, encoding='utf-8-sig')
    today_str = datetime.date.today().strftime("%Y-%m-%d")
    today_df = current_logs[current_logs["날짜"] == today_str]
    
    if not today_df.empty:
        display_df = today_df.copy()
        display_df.insert(0, '기록번호', today_df.index)
        st.dataframe(display_df, use_container_width=True, hide_index=True)
        
        with st.expander("🛠️ 방금 입력한 기록 수정 / 삭제하기", expanded=False):
            options = today_df.index.tolist()
            def make_label(idx):
                r = today_df.loc[idx]
                utg_str = f" | 유통기한: {r['유통기한']}" if pd.notna(r['유통기한']) and str(r['유통기한']) != 'nan' and r['유통기한'] != '' else ""
                return f"번호[{idx}] - {r['품목명']} ({r['구분']} {r['수량']}개{utg_str})"
                
            col_m1, col_m2, col_m3 = st.columns([2, 1, 2])
            with col_m1:
                selected_idx = st.selectbox("수정/삭제할 기록 줄 선택", options, format_func=make_label)
            with col_m2:
                manage_action = st.radio("작업 선택", ["수량 수정", "기록 삭제"])
            with col_m3:
                current_target = today_df.loc[selected_idx]
                if manage_action == "수량 수정":
                    new_qty = st.number_input("변경할 새 수량 입력", min_value=1, step=1, value=int(current_target['수량']))
                    execute_btn = st.button("🔧 수량 수정 완료")
                else:
                    st.write(f"⚠️ 정말로 이 기록을 삭제하시겠습니까?")
                    execute_btn = st.button("❌ 선택 기록 삭제 확정")
                    
            if execute_btn:
                if manage_action == "기록 삭제":
                    updated_logs = current_logs.drop(selected_idx).reset_index(drop=True)
                    st.session_state.success_msg = "❌ 선택하신 내역이 정상적으로 삭제되었습니다."
                elif manage_action == "수량 수정":
                    current_logs.loc[selected_idx, '수량'] = new_qty
                    updated_logs = current_logs
                    st.session_state.success_msg = f"🔧 수량이 {new_qty}개로 정상 수정되었습니다."
                
                updated_logs.to_csv(target_log_file, index=False, encoding='utf-8-sig')
                st.cache_data.clear()
                st.rerun()
    else:
        st.caption("오늘 아직 입력된 내역이 없습니다.")

# 1) 입고 등록 메뉴(개별)
if menu == "📥 물품 입고 등록 (개별)":
    st.subheader("📥 매장 물품 입고 등록 (개별)")
    st.info("💡 입고 수량และ 유통기한을 맞춘 후 **엔터(Enter) 키**를 누르면 즉시 저장 및 비워집니다.")
    
    categories = list(master_data['대분류'].unique()) if not master_data.empty else ["원재료", "부자재", "디저트&완제품"]
    selected_cat = st.selectbox("1. 입고 품목 분류 선택", categories, key="in_cat")
    
    filtered_items = master_data[master_data['대분류'] == selected_cat]
    item_list = filtered_items['품목명'].drop_duplicates().tolist()
    
    if not item_list:
        st.warning("등록된 품목이 없습니다.")
    else:
        with st.form("inbound_form"):
            col1, col2 = st.columns(2)
            with col1:
                st.date_input("날짜 선택", datetime.date.today(), key="in_date")
                selected_item = st.selectbox("2. 입고 품목 선택", item_list, key="in_item")
            with col2:
                item_info = filtered_items[filtered_items['품목명'] == selected_item].iloc[0]
                existing_utg = str(item_info['유통기한']).strip()
                
                default_utg_date = datetime.date.today()
                if existing_utg and existing_utg != '0' and existing_utg != '1899-12-31' and existing_utg != 'nan' and existing_utg != '':
                    try: default_utg_date = pd.to_datetime(existing_utg).date()
                    except: pass
                    
                st.caption(f"📊 기존 정보 ➡️ [현재 엑셀상 기본재고: {item_info['엑셀기본재고']}개 / 유통기한: {existing_utg}]")
                st.date_input("3. 유통기한 확인 및 변경 설정", default_utg_date, key="in_utg")
                
            st.number_input("4. 입고 수량 입력 후 엔터(Enter)", min_value=0, step=1, key="in_qty")
            st.form_submit_button("📥 입고 데이터 저장하기", on_click=save_inbound_callback)
                    
        show_today_logs_and_management()

# 2) 출고 등록 메뉴(개별)
elif menu == "📤 물품 출고 등록 (소모-개별)":
    st.subheader("📤 매장 소모(출고) 등록 (개별)")
    st.info("💡 출고(소모) 수량을 입력한 뒤 **엔터(Enter) 키**를 누르면 즉시 저장 및 비워집니다.")
    
    categories = list(master_data['대분류'].unique()) if not master_data.empty else ["원재료", "부자재", "디저트&완제품"]
    selected_cat = st.selectbox("1. 출고 품목 분류 선택", categories, key="out_cat")
    
    filtered_items = master_data[master_data['대분류'] == selected_cat]
    item_list = filtered_items['품목명'].drop_duplicates().tolist()
    
    if not item_list:
        st.warning("등록된 품목이 없습니다.")
    else:
        with st.form("outbound_form"):
            col1, col2 = st.columns(2)
            with col1:
                st.date_input("날짜 선택", datetime.date.today(), key="out_date")
                selected_item = st.selectbox("2. 출고 품목 선택", item_list, key="out_item")
            with col2:
                item_info = filtered_items[filtered_items['품목명'] == selected_item].iloc[0]
                st.caption(f"📊 현재 재고 참고 ➡️ [엑셀상 기본재고: {item_info['엑셀기본재고']}개 / 유통기한: {item_info['유통기한']}]")
                
            st.number_input("3. 소모(출고) 수량 입력 후 엔터(Enter)", min_value=0, step=1, key="out_qty")
            st.form_submit_button("📤 출고 데이터 저장하기", on_click=save_outbound_callback)
                    
        show_today_logs_and_management()

# 3) 전품목 일괄 입력 (엑셀 스타일) -> [수정 반영 완료]
elif menu == "📝 전품목 일괄 입력 (엑셀 스타일)":
    st.subheader("📝 전품목 일괄 입력 (엑셀 스타일)")
    
    if st.session_state.bulk_download_ready and st.session_state.bulk_excel_bytes:
        st.success("🎉 현재 상태를 반영한 수불 집계표 및 원본 대장 파일 추출이 완료되었습니다!")
        col_dl1, col_dl2 = st.columns(2)
        
        with col_dl1:
            st.download_button(
                label="📊 방금 저장된 월간 요약 집계표 다운로드 (Excel)",
                data=st.session_state.bulk_excel_bytes,
                file_name=st.session_state.bulk_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                key="vini_bulk_summary_fixed_key"
            )
        with col_dl2:
            if st.session_state.orig_excel_bytes:
                st.download_button(
                    label="📥 수정된 서버 원본 엑셀 전체 파일 가져오기 (.xlsx)",
                    data=st.session_state.orig_excel_bytes,
                    file_name=f"🟢수치최신반영_VINI대장_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="vini_orig_master_fixed_key"
                )
        st.markdown("---")
    
    col_t1, col_t2 = st.columns([1, 2])
    with col_t1:
        bulk_date = st.date_input("🗓️ 기록할 날짜 선택", datetime.date.today(), key="bulk_date")
    with col_t2:
        bulk_cat = st.radio("분류 필터링", ["전체"] + list(master_data['대분류'].unique()), horizontal=True, key="bulk_cat")
        
    current_logs = pd.read_csv(STOCK_LOG_FILE, encoding='utf-8-sig')
    
    if not current_logs.empty:
        pivot_all = current_logs.pivot_table(
            index=['대분류', '품목명'], columns='구분', values='수량', aggfunc='sum'
        ).fillna(0).reset_index()
        if "금월 입고" not in pivot_all.columns: pivot_all["금월 입고"] = 0
        if "월 소모(출고)" not in pivot_all.columns: pivot_all["월 소모(출고)"] = 0
    else:
        pivot_all = pd.DataFrame(columns=['대분류', '품목명', '금월 입고', '월 소모(출고)'])
        
    bulk_df = pd.merge(master_data[['대분류', '품목명', '엑셀기본재고', '유통기한']], pivot_all, on=['대분류', '품목명'], how='left').fillna(0)
    bulk_df['currently_stock'] = (bulk_df['엑셀기본재고'] + bulk_df['금월 입고'] - bulk_df['월 소모(출고)']).astype(int)
    
    bulk_df['📥 오늘 입고량'] = 0
    bulk_df['📤 오늘 소모량'] = 0
    
    bulk_df = bulk_df.rename(columns={"유통기한": "⏳ 유통기한", "currently_stock": "현재고"})
    display_bulk = bulk_df[['대분류', '품목명', '현재고', '📥 오늘 입고량', '📤 오늘 소모량', '⏳ 유통기한']].copy()
    
    if bulk_cat != "전체":
        display_bulk = display_bulk[display_bulk['대분류'] == bulk_cat]
        
    edited_bulk = st.data_editor(
        display_bulk,
        column_config={
            "대분류": st.column_config.TextColumn("분류", disabled=True),
            "품목명": st.column_config.TextColumn("품목명", disabled=True),
            "현재고": st.column_config.NumberColumn("현재고 (참고용)", disabled=True, format="%d"),
            "📥 오늘 입고량": st.column_config.NumberColumn("오늘 입고량 입력", min_value=0, step=1, format="%d"),
            "📤 오늘 소모량": st.column_config.NumberColumn("오늘 소모량 입력", min_value=0, step=1, format="%d"),
            "⏳ 유통기한": st.column_config.TextColumn("유통기한 입력 (변경 시 수정 가능)"),
        },
        use_container_width=True,
        hide_index=True,
        key="bulk_data_editor"
    )
    
    if st.button("💾 위 입력된 모든 내역 일괄 저장하고 엑셀 추출하기", use_container_width=True):
        batch_new_logs = []
        master_update_needed = False
        
        if os.path.exists(ORIGINAL_EXCEL_PATH):
            try:
                wb = openpyxl.load_workbook(ORIGINAL_EXCEL_PATH)
                target_day_text = f"{bulk_date.day}일"
                target_day_num = str(bulk_date.day)
                
                for idx, row in edited_bulk.iterrows():
                    p_name = row['품목명']
                    p_cat = row['대분류']
                    
                    try: in_val = int(float(str(row['📥 오늘 입고량']).strip() or 0))
                    except: in_val = 0
                        
                    try: out_val = int(float(str(row['📤 오늘 소모량']).strip() or 0))
                    except: out_val = 0
                        
                    utg_val = str(row['⏳ 유통기한']).strip() if pd.notna(row['⏳ 유통기한']) else ""
                    
                    target_sheet = SHEET_MAP.get(p_cat, p_cat) 
                    if target_sheet in wb.sheetnames:
                        ws = wb[target_sheet]
                        header_row = 3
                        
                        name_idx, inbound_idx, target_date_col_idx, expiry_idx = None, None, None, None
                        for c in range(1, ws.max_column + 1):
                            cell_v = ws.cell(row=header_row, column=c).value
                            if cell_v is None: continue
                            val = str(cell_v).strip().replace(" ", "")
                            
                            if '품목이름' in val or '품목명' in val or '구분' in val: name_idx = c
                            elif '금월입고' in val or '입고' in val: inbound_idx = c 
                            elif val == target_day_text or val == target_day_num or val == f"0{target_day_num}": target_date_col_idx = c 
                            elif '유통' in val: expiry_idx = c

                        if name_idx:
                            for r in range(header_row + 1, ws.max_row + 1):
                                excel_item_name = str(ws.cell(row=r, column=name_idx).value).strip()
                                if excel_item_name == p_name:
                                    if inbound_idx and in_val > 0:
                                        cell_in = ws.cell(row=r, column=inbound_idx)
                                        try: current_in = int(cell_in.value) if cell_in.value is not None else 0
                                        except: current_in = 0
                                        cell_in.value = current_in + in_val
                                        
                                    if target_date_col_idx and out_val > 0:
                                        cell_out = ws.cell(row=r, column=target_date_col_idx)
                                        try: current_out = int(cell_out.value) if cell_out.value is not None else 0
                                        except: current_out = 0
                                        cell_out.value = current_out + out_val
                                        
                                    if expiry_idx and utg_val:
                                        ws.cell(row=r, column=expiry_idx).value = utg_val
                                    break
                wb.save(ORIGINAL_EXCEL_PATH)
                wb.close()
            except Exception as e:
                st.error(f"서버 엑셀 일괄 규칙 반영 실패: {e}")

        for idx, row in edited_bulk.iterrows():
            p_name = row['품목명']
            p_cat = row['대분류']
            
            try: in_val = int(float(str(row['📥 오늘 입고량']).strip() or 0))
            except: in_val = 0
                
            try: out_val = int(float(str(row['📤 오늘 소모량']).strip() or 0))
            except: out_val = 0
                
            utg_val = str(row['⏳ 유통기한']).strip() if pd.notna(row['⏳ 유통기한']) else ""
            
            orig_match = master_data[master_data['품목명'] == p_name]
            if not orig_match.empty:
                orig_utg = str(orig_match.iloc[0]['유통기한']).strip()
                if utg_val != orig_utg:
                    master_data.loc[master_data['품목명'] == p_name, '유통기한'] = utg_val
                    master_update_needed = True
            
            if in_val > 0:
                batch_new_logs.append({"날짜": bulk_date.strftime("%Y-%m-%d"), "대분류": p_cat, "품목명": p_name, "구분": "금월 입고", "수량": in_val, "유통기한": utg_val})
            if out_val > 0:
                batch_new_logs.append({"날짜": bulk_date.strftime("%Y-%m-%d"), "대분류": p_cat, "품목명": p_name, "구분": "월 소모(출고)", "수량": out_val, "유통기한": ""})
                
        if batch_new_logs:
            existing_logs = pd.read_csv(STOCK_LOG_FILE, encoding='utf-8-sig')
            combined_logs = pd.concat([existing_logs, pd.DataFrame(batch_new_logs)], ignore_index=True)
            combined_logs.to_csv(STOCK_LOG_FILE, index=False, encoding='utf-8-sig')
            
        if master_update_needed:
            master_data.to_csv(CUSTOM_MASTER_FILE, index=False, encoding='utf-8-sig')
            
        # 변동 수치 유무 관계없이 무조건 다운로드용 버퍼를 새로 생성하여 세션 주입
        fresh_master, fresh_logs = get_latest_master_and_logs()
        current_month_str = bulk_date.strftime("%Y-%m")
        
        if not fresh_logs.empty:
            fresh_logs['날짜_dt'] = pd.to_datetime(fresh_logs['날짜'])
            filtered_df = fresh_logs[fresh_logs['날짜_dt'].dt.strftime("%Y-%m") == current_month_str]
        else:
            filtered_df = pd.DataFrame()
            
        if not filtered_df.empty:
            summary = filtered_df.pivot_table(index=['대분류', '품목명'], columns='구분', values='수량', aggfunc='sum').fillna(0).reset_index()
        else:
            summary = pd.DataFrame(columns=['대분류', '품목명', '금월 입고', '월 소모(출고)'])
            
        if "금월 입고" not in summary.columns: summary["금월 입고"] = 0
        if "월 소모(출고)" not in summary.columns: summary["월 소모(출고)"] = 0
        
        summary['금월 입고'] = summary['금월 입고'].astype(int)
        summary['월 소모(출고)'] = summary['월 소모(출고)'].astype(int)
        summary = summary.rename(columns={"금월 입고": "총 입고량", "월 소모(출고)": "총 소모량"})
        summary = pd.merge(fresh_master[['대분류', '품목명', '엑셀기본재고', '유통기한']], summary, on=['대분류', '품목명'], how='left').fillna(0)
        summary['실시간 예상 현재고'] = summary['엑셀기본재고'] + summary['총 입고량'] - summary['총 소모량']
        
        summary['총 입고량'] = summary['총 입고량'].astype(int)
        summary['총 소모량'] = summary['총 소모량'].astype(int)
        summary['실시간 예상 현재고'] = summary['실시간 예상 현재고'].astype(int)
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            summary.to_excel(writer, index=False, sheet_name=f'{current_month_str}_일괄갱신집계')
        
        st.session_state.bulk_excel_bytes = buffer.getvalue()
        st.session_state.bulk_filename = f"vini_bulk_updated_{bulk_date.strftime('%Y%m%d')}.xlsx"
        st.session_state.bulk_download_ready = True
        
        if os.path.exists(ORIGINAL_EXCEL_PATH):
            with open(ORIGINAL_EXCEL_PATH, "rb") as f:
                st.session_state.orig_excel_bytes = f.read()
        
        st.cache_data.clear()
        st.rerun()

    show_today_logs_and_management()

# 🍷 4) 와인 재고 관리 메뉴 (일괄 입력 엑셀 스타일) -> [수정 반영 완료]
elif menu == "🍷 와인 재고 관리 (일괄 입력)":
    st.subheader("🍷 와인 수불 대장 관리 및 일괄 입력 (엑셀 스타일)")
    st.info("💡 와인 전용 대장인 `와인_입출고양식_디자인적용_현재고요약.xlsx` 파일을 타겟으로 동기화합니다.")
    
    if st.session_state.wine_download_ready and st.session_state.wine_excel_bytes:
        st.success("🎉 현재 상태를 반영한 와인 대장 파일 추출이 완료되었습니다!")
        col_wd1, col_wd2 = st.columns(2)
        
        with col_wd1:
            st.download_button(
                label="📊 방금 저장된 와인 수불 집계표 다운로드 (Excel)",
                data=st.session_state.wine_excel_bytes,
                file_name=st.session_state.wine_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                key="wine_bulk_summary_fixed_key"
            )
        with col_wd2:
            if st.session_state.wine_orig_bytes:
                st.download_button(
                    label="📥 수정된 와인 대장 전체 원본 파일 가져오기 (.xlsx)",
                    data=st.session_state.wine_orig_bytes,
                    file_name=f"🟢수치최신반영_와인대장_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="wine_orig_master_fixed_key"
                )
        st.markdown("---")
        
    w_bulk_date = st.date_input("🗓️ 기록할 날짜 선택", datetime.date.today(), key="wine_bulk_date")
    
    current_wine_logs = pd.read_csv(WINE_LOG_FILE, encoding='utf-8-sig')
    
    if not current_wine_logs.empty:
        w_pivot = current_wine_logs.pivot_table(
            index=['대분류', '품목명'], columns='구분', values='수량', aggfunc='sum'
        ).fillna(0).reset_index()
        if "금월 입고" not in w_pivot.columns: w_pivot["금월 입고"] = 0
        if "월 소모(출고)" not in w_pivot.columns: w_pivot["월 소모(출고)"] = 0
    else:
        w_pivot = pd.DataFrame(columns=['대분류', '품목명', '금월 입고', '월 소모(출고)'])
        
    w_bulk_df = pd.merge(wine_master_data[['대분류', '품목명', '엑셀기본재고', '유통기한']], w_pivot, on=['대분류', '품목명'], how='left').fillna(0)
    w_bulk_df['currently_stock'] = (w_bulk_df['엑셀기본재고'] + w_bulk_df['금월 입고'] - w_bulk_df['월 소모(출고)']).astype(int)
    
    w_bulk_df['📥 오늘 입고량'] = 0
    w_bulk_df['📤 오늘 소모량'] = 0
    
    w_bulk_df = w_bulk_df.rename(columns={"유통기한": "⏳ 빈티지/유통기한", "currently_stock": "현재고"})
    w_display_bulk = w_bulk_df[['대분류', '품목명', '현재고', '📥 오늘 입고량', '📤 오늘 소모량', '⏳ 빈티지/유통기한']].copy()
    
    edited_wine_bulk = st.data_editor(
        w_display_bulk,
        column_config={
            "대분류": st.column_config.TextColumn("분류", disabled=True),
            "품목명": st.column_config.TextColumn("와인 품목명", disabled=True),
            "현재고": st.column_config.NumberColumn("현재고 (참고용)", disabled=True, format="%d"),
            "📥 오늘 입고량": st.column_config.NumberColumn("오늘 입고량 입력", min_value=0, step=1, format="%d"),
            "📤 오늘 소모량": st.column_config.NumberColumn("오늘 소모량 입력", min_value=0, step=1, format="%d"),
            "⏳ 빈티지/유통기한": st.column_config.TextColumn("정보 수정"),
        },
        use_container_width=True,
        hide_index=True,
        key="wine_data_editor"
    )
    
    if st.button("💾 위 입력된 와인 내역 일괄 저장하고 엑셀 추출하기", use_container_width=True):
        w_new_logs = []
        w_master_update = False
        
        if os.path.exists(WINE_EXCEL_PATH):
            try:
                wb = openpyxl.load_workbook(WINE_EXCEL_PATH)
                target_day_text = f"{w_bulk_date.day}일"
                target_day_num = str(w_bulk_date.day)
                
                target_sheet = wb.sheetnames[0]
                for name in wb.sheetnames:
                    if "와인" in name or "재고" in name:
                        target_sheet = name
                        break
                        
                ws = wb[target_sheet]
                header_row = 3
                
                name_idx, inbound_idx, target_date_col_idx, expiry_idx = None, None, None, None
                for c in range(1, ws.max_column + 1):
                    cell_v = ws.cell(row=header_row, column=c).value
                    if cell_v is None: continue
                    val = str(cell_v).strip().replace(" ", "")
                    
                    if '품목' in val or '와인' in val or '이름' in val or '구분' in val: name_idx = c
                    elif '입고' in val: inbound_idx = c
                    elif val == target_day_text or val == target_day_num or val == f"0{target_day_num}": target_date_col_idx = c
                    elif '유통' in val or '빈티지' in val: expiry_idx = c
                    
                if name_idx:
                    for idx, row in edited_wine_bulk.iterrows():
                        p_name = row['품목명']
                        try: in_val = int(float(str(row['📥 오늘 입고량']).strip() or 0))
                        except: in_val = 0
                        try: out_val = int(float(str(row['📤 오늘 소모량']).strip() or 0))
                        except: out_val = 0
                        utg_val = str(row['⏳ 빈티지/유통기한']).strip() if pd.notna(row['⏳ 빈티지/유통기한']) else ""
                        
                        for r in range(header_row + 1, ws.max_row + 1):
                            ex_item = str(ws.cell(row=r, column=name_idx).value).strip()
                            if ex_item == p_name:
                                if inbound_idx and in_val > 0:
                                    cell_in = ws.cell(row=r, column=inbound_idx)
                                    cur_in = int(cell_in.value) if cell_in.value is not None else 0
                                    cell_in.value = cur_in + in_val
                                if target_date_col_idx and out_val > 0:
                                    cell_out = ws.cell(row=r, column=target_date_col_idx)
                                    cur_out = int(cell_out.value) if cell_out.value is not None else 0
                                    cell_out.value = cur_out + out_val
                                if expiry_idx and utg_val:
                                    ws.cell(row=r, column=expiry_idx).value = utg_val
                                break
                                
                wb.save(WINE_EXCEL_PATH)
                wb.close()
            except Exception as e:
                st.error(f"와인 엑셀 반영 중 오류 발생: {e}")
                
        for idx, row in edited_wine_bulk.iterrows():
            p_name = row['품목명']
            try: in_val = int(float(str(row['📥 오늘 입고량']).strip() or 0))
            except: in_val = 0
            try: out_val = int(float(str(row['📤 오늘 소모량']).strip() or 0))
            except: out_val = 0
            utg_val = str(row['⏳ 빈티지/유통기한']).strip() if pd.notna(row['⏳ 빈티지/유통기한']) else ""
            
            orig_match = wine_master_data[wine_master_data['품목명'] == p_name]
            if not orig_match.empty:
                if utg_val != str(orig_match.iloc[0]['유통기한']).strip():
                    wine_master_data.loc[wine_master_data['품목명'] == p_name, '유통기한'] = utg_val
                    w_master_update = True
                    
            if in_val > 0:
                w_new_logs.append({"날짜": w_bulk_date.strftime("%Y-%m-%d"), "대분류": "와인", "품목명": p_name, "구분": "금월 입고", "수량": in_val, "유통기한": utg_val})
            if out_val > 0:
                w_new_logs.append({"날짜": w_bulk_date.strftime("%Y-%m-%d"), "대분류": "와인", "품목명": p_name, "구분": "월 소모(출고)", "수량": out_val, "유통기한": ""})
                
        if w_new_logs:
            ex_logs = pd.read_csv(WINE_LOG_FILE, encoding='utf-8-sig')
            pd.concat([ex_logs, pd.DataFrame(w_new_logs)], ignore_index=True).to_csv(WINE_LOG_FILE, index=False, encoding='utf-8-sig')
        if w_master_update:
            wine_master_data.to_csv(WINE_MASTER_FILE, index=False, encoding='utf-8-sig')
            
        # 와인 변동 수치 유무 관계없이 무조건 와인 엑셀 데이터를 바이트로 로드 및 세션 주입
        fresh_w_master, fresh_w_logs = get_wine_master_and_logs()
        w_month = w_bulk_date.strftime("%Y-%m")
        
        if not fresh_w_logs.empty:
            fresh_w_logs['날짜_dt'] = pd.to_datetime(fresh_w_logs['날짜'])
            w_fil = fresh_w_logs[fresh_w_logs['날짜_dt'].dt.strftime("%Y-%m") == w_month]
        else:
            w_fil = pd.DataFrame()
            
        if not w_fil.empty:
            w_sum = w_fil.pivot_table(index=['대분류', '품목명'], columns='구분', values='수량', aggfunc='sum').fillna(0).reset_index()
        else:
            w_sum = pd.DataFrame(columns=['대분류', '품목명', '금월 입고', '월 소모(출고)'])
            
        if "금월 입고" not in w_sum.columns: w_sum["금월 입고"] = 0
        if "월 소모(출고)" not in w_sum.columns: w_sum["월 소모(출고)"] = 0
        
        w_sum = w_sum.rename(columns={"금월 입고": "총 입고량", "월 소모(출고)": "총 소모량"})
        w_sum = pd.merge(fresh_w_master[['대분류', '품목명', '엑셀기본재고', '유통기한']], w_sum, on=['대분류', '품목명'], how='left').fillna(0)
        w_sum['실시간 예상 현재고'] = w_sum['엑셀기본재고'] + w_sum['총 입고량'] - w_sum['총 소모량']
        
        w_sum['총 입고량'] = w_sum['총 입고량'].astype(int)
        w_sum['총 소모량'] = w_sum['총 소모량'].astype(int)
        w_sum['실시간 예상 현재고'] = w_sum['실시간 예상 현재고'].astype(int)
        
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            w_sum.to_excel(writer, index=False, sheet_name=f'{w_month}_와인수불집계')
        st.session_state.wine_excel_bytes = buf.getvalue()
        st.session_state.wine_filename = f"wine_summary_{w_bulk_date.strftime('%Y%m%d')}.xlsx"
        st.session_state.wine_download_ready = True
        
        if os.path.exists(WINE_EXCEL_PATH):
            with open(WINE_EXCEL_PATH, "rb") as f:
                st.session_state.wine_orig_bytes = f.read()
        st.cache_data.clear()
        st.rerun()
        
    show_today_logs_and_management(is_wine=True)

# 5) 실시간 현재고 현황판
elif menu == "📋 실시간 현재고 현황판":
    st.subheader("📋 매장 실시간 현재고 현황판")
    
    fresh_master, fresh_logs = get_latest_master_and_logs()
    
    if not fresh_logs.empty:
        pivot_all = fresh_logs.pivot_table(
            index=['대분류', '품목명'], columns='구분', values='수량', aggfunc='sum'
        ).fillna(0).reset_index()
        if "금월 입고" not in pivot_all.columns: pivot_all["금월 입고"] = 0
        if "월 소모(출고)" not in pivot_all.columns: pivot_all["월 소모(출고)"] = 0
    else:
        pivot_all = pd.DataFrame(columns=['대분류', '품목명', '금월 입고', '월 소모(출고)'])
        
    dashboard_df = pd.merge(fresh_master[['대분류', '품목명', '엑셀기본재고', '유통기한']], pivot_all, on=['대분류', '품목명'], how='left').fillna(0)
    
    dashboard_df['금월 입고'] = dashboard_df['금월 입고'].astype(int)
    dashboard_df['월 소모(출고)'] = dashboard_df['월 소모(출고)'].astype(int)
    dashboard_df['currently_stock'] = dashboard_df['엑셀기본재고'] + dashboard_df['금월 입고'] - dashboard_df['월 소모(출고)']
    dashboard_df['currently_stock'] = dashboard_df['currently_stock'].astype(int)
    
    dashboard_df = dashboard_df.rename(columns={"엑셀기본재고": "기본재고(이월)", "금월 입고": "누적 입고량", "월 소모(출고)": "누적 소모량", "currently_stock": "현재고"})
    
    col_f1, col_f2 = st.columns([1, 2])
    with col_f1: filter_cat = st.radio("분류별 필터", ["전체"] + list(fresh_master['대분류'].unique()), horizontal=True)
    with col_f2: search_query = st.text_input("🔍 품목 실시간 키워드 검색", "")
        
    display_dash = dashboard_df.copy()
    if filter_cat != "전체": display_dash = display_dash[display_dash['대분류'] == filter_cat]
    if search_query: display_dash = display_dash[display_dash['품목명'].str.contains(search_query, case=False)]
        
    display_dash = display_dash.sort_values(by="현재고", ascending=True)
    
    def highlight_shortage(row):
        styles = [''] * len(row)
        if row['현재고'] <= 0:
            idx = row.index.get_loc('현재고')
            styles[idx] = 'background-color: #FADBD8; color: #78281F; font-weight: bold;'
        return styles

    if not display_dash.empty:
        styled_dash = display_dash.style.apply(highlight_shortage, axis=1)
        st.dataframe(styled_dash, use_container_width=True, hide_index=True)

# 6) 월별 조회 및 백업 화면
elif menu == "📊 월별 수불 대장 및 백업 다운로드":
    st.subheader("월별 수불 대장 및 데이터 다운로드")
    fresh_master, fresh_logs = get_latest_master_and_logs()
    
    if fresh_logs.empty:
        st.info("아직 누적된 데이터가 없습니다.")
    else:
        fresh_logs['날짜'] = pd.to_datetime(fresh_logs['날짜'])
        fresh_logs['년월'] = fresh_logs['날짜'].dt.to_period('M').astype(str)
        
        col1, col2 = st.columns(2)
        with col1:
            available_months = sorted(fresh_logs['년월'].unique(), reverse=True)
            selected_month = st.selectbox("조회할 월 선택", available_months)
        with col2:
            selected_cat = st.selectbox("분류 필터", ["전체"] + list(fresh_master['대분류'].unique()))
            
        filtered_df = fresh_logs[fresh_logs['년월'] == selected_month]
        if selected_cat != "전체":
            filtered_df = filtered_df[filtered_df['대분류'] == selected_cat]
            
        if filtered_df.empty:
            st.warning("선택한 조건에 맞는 기록이 없습니다.")
        else:
            summary = filtered_df.pivot_table(
                index=['대분류', '품목명'], columns='구분', values='수량', aggfunc='sum'
            ).fillna(0).reset_index()
            
            if "금월 입고" not in summary.columns: summary["금월 입고"] = 0
            if "월 소모(출고)" not in summary.columns: summary["월 소모(출고)"] = 0
            
            summary['금월 입고'] = summary['금월 입고'].astype(int)
            summary['월 소모(출고)'] = summary['월 소모(출고)'].astype(int)
            summary = summary.rename(columns={"금월 입고": "총 입고량", "월 소모(출고)": "총 소모량"})
            summary = pd.merge(fresh_master[['대분류', '품목명', '엑셀기본재고', '유통기한']], summary, on=['대분류', '품목명'], how='inner')
            summary['실시간 예상 현재고'] = summary['엑셀기본재고'] + summary['총 입고량'] - summary['총 소모량']
            
            st.markdown(f"### 📅 {selected_month} 품목별 종합 수불 집계")
            st.dataframe(summary, use_container_width=True, hide_index=True)
            
            st.markdown("---")
            st.subheader("💾 데이터 안전 백업 및 내보내기 (Excel)")
            
            buffer_summary = io.BytesIO()
            with pd.ExcelWriter(buffer_summary, engine='openpyxl') as writer:
                summary.to_excel(writer, index=False, sheet_name=f'{selected_month}_수불집계')
            excel_summary_bytes = buffer_summary.getvalue()
            
            full_raw_logs = pd.read_csv(STOCK_LOG_FILE, encoding='utf-8-sig')
            buffer_raw = io.BytesIO()
            with pd.ExcelWriter(buffer_raw, engine='openpyxl') as writer:
                full_raw_logs.to_excel(writer, index=False, sheet_name='전체누적로그')
            excel_raw_bytes = buffer_raw.getvalue()
            
            col_b1, col_b2 = st.columns(2)
            with col_b1:
                st.download_button(
                    label=f"📊 {selected_month} 월간 집계표 다운로드 (Excel)",
                    data=excel_summary_bytes,
                    file_name=f"vini_coffee_summary_{selected_month}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="monthly_summary_download_btn"
                )
            with col_b2:
                st.download_button(
                    label="📝 전체 일별 로그 백업 다운로드 (Excel)",
                    data=excel_raw_bytes,
                    file_name="vini_daily_stock_log_backup.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="daily_log_backup_download_btn"
                )

# 7) 품목 추가 / 삭제 관리 화면
elif menu == "⚙️ 품목 추가/삭제 관리":
    st.subheader("⚙️ 매장 품목 추가 및 삭제 관리")
    
    col_a, col_b = st.columns(2)
    
    with col_a:
        st.markdown("### 📥 신규 품목 추가")
        with st.form("add_item_form", clear_on_submit=True):
            add_cat = st.selectbox("품목 분류 선택", ["원재료", "부자재", "디저트&완제품"])
            add_name = st.text_input("새로운 품목명 입력 (정확하게)")
            add_stock = st.number_input("초기 기본 이월 재고량 (개수)", min_value=0, step=1, value=0)
            add_utg = st.date_input("품목 기본 유통기한 지정", datetime.date.today())
            
            submit_add = st.form_submit_button("➕ 마스터 품목 추가 완료")
            if submit_add:
                if not add_name.strip():
                    st.error("⚠️ 품목 이름을 공백으로 추가할 수 없습니다.")
                elif add_name.strip() in master_data['품목명'].values:
                    st.warning(f"⚠️ 이미 존재하는 품목명입니다: {add_name}")
                else:
                    new_row = pd.DataFrame([{
                        "품목명": add_name.strip(),
                        "대분류": add_cat,
                        "유통기한": add_utg.strftime("%Y-%m-%d"),
                        "엑셀기본재고": int(add_stock)
                    }])
                    master_data = pd.concat([master_data, new_row], ignore_index=True)
                    master_data.to_csv(CUSTOM_MASTER_FILE, index=False, encoding='utf-8-sig')
                    st.session_state.success_msg = f"✅ 신규 품목 등록 성공: [{add_cat}] {add_name}"
                    st.cache_data.clear()
                    st.rerun()

    with col_b:
        st.markdown("### ❌ 기존 품목 삭제 (단종)")
        if master_data.empty:
            st.caption("삭제할 품목이 마스터에 존재하지 않습니다.")
        else:
            all_current_items = sorted(master_data['품목명'].tolist())
            target_delete_item = st.selectbox("시스템에서 완전히 제거할 품목 선택", all_current_items)
            submit_delete = st.button("❌ 선택 품목 영구 삭제 확정")
            
            if submit_delete:
                master_data = master_data[master_data['품목명'] != target_delete_item]
                master_data.to_csv(CUSTOM_MASTER_FILE, index=False, encoding='utf-8-sig')
                st.session_state.success_msg = f"❌ 품목 삭제 완료: {target_delete_item}"
                st.cache_data.clear()
                st.rerun()
                
    st.markdown("---")
    st.dataframe(master_data.sort_values(by=["대분류", "품목명"]), use_container_width=True, hide_index=True)


# 8) 데이터 관리 및 엑셀 동기화 통합 메뉴
elif menu == "🛠️ 데이터 관리 및 엑셀 동기화":
    st.subheader("🛠️ 시스템 데이터 관리 및 엑셀 파일 동기화")
    st.markdown("---")
    
    st.markdown("### 📅 새 엑셀파일 업로드 및 실시간 품목 동기화")
    st.info("💡 VINI COFFEE 매출관리 시스템 대장(.xlsx) 파일을 업로드하고 버튼을 누르면 원재료, 부자재, 완제품의 재고 현황을 자동으로 파싱하여 동기화합니다.")
    
    uploaded_file = st.file_uploader("엑셀 대장 파일을 업로드하세요 (.xlsx)", type=["xlsx"], key="vini_uploader")

    if uploaded_file is not None:
        st.success("파일이 웹 브라우저에 성공적으로 준비되었습니다!")
        
        if st.button("🚀 업로드된 엑셀 데이터 파싱 및 연동 시작", use_container_width=True):
            try:
                xl = pd.ExcelFile(uploaded_file)
                sheets_to_try = {
                    "원재료": ["원재료", "원재료(간략)"],
                    "부자재": ["부자재", "부자재(간략)"],
                    "디저트&완제품": ["디저트&완제품", "디저트&완제품(간략)"]
                }
                
                uploaded_master_list = []
                
                for cat, standard_names in sheets_to_try.items():
                    target_sheet = None
                    for name in xl.sheet_names:
                        if name.strip() in standard_names or cat in name:
                            target_sheet = name
                            break
                    
                    if target_sheet:
                        df = pd.read_excel(xl, sheet_name=target_sheet, skiprows=2)
                        df.columns = [str(c).strip().replace(" ", "") for c in df.columns]
                        
                        name_col = None
                        for col in df.columns:
                            if '품목이름' in col or '구분' in col or '품목명' in col:
                                name_col = col
                                break
                        
                        if name_col:
                            df = df.dropna(subset=[name_col])
                            df = df[df[name_col].astype(str).str.strip() != '0']
                            df = df[df[name_col].astype(str).str.strip() != '']
                            
                            expiry_col = [c for c in df.columns if '유통' in c]
                            stock_col = [c for c in df.columns if '재고' in c]
                            
                            temp_df = pd.DataFrame()
                            temp_df['품목명'] = df[name_col].astype(str).str.strip()
                            temp_df['대분류'] = cat
                            
                            if expiry_col:
                                temp_df['유통기한'] = df[expiry_col[0]].astype(str).str.strip()
                            else:
                                temp_df['유통기한'] = ""
                                
                            if stock_col:
                                temp_df['엑셀기본재고'] = pd.to_numeric(df[stock_col[0]], errors='coerce').fillna(0).astype(int)
                            else:
                                temp_df['엑셀기본재고'] = 0
                                
                            uploaded_master_list.append(temp_df)
                
                if uploaded_master_list:
                    final_uploaded_master = pd.concat(uploaded_master_list, ignore_index=True)
                    final_uploaded_master.to_csv(CUSTOM_MASTER_FILE, index=False, encoding='utf-8-sig')
                    
                    with open(ORIGINAL_EXCEL_PATH, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                        
                    st.subheader("📋 새 엑셀파일에서 정상 인식된 대장 내역")
                    st.dataframe(final_uploaded_master, use_container_width=True, hide_index=True)
                    st.balloons()
                    st.success("🎉 업로드된 엑셀 파일의 시트별 품목과 재고량이 시스템 마스터에 오차 없이 완벽 연동되었습니다!")
                    st.cache_data.clear()
                else:
                    st.error("⚠️ 업로드된 엑셀 파일에서 '원재료', '부자재' 또는 '디저트' 시트를 매칭하지 못했거나 품목 열을 찾을 수 없습니다.")
                    
            except Exception as e:
                st.error(f"엑셀 대장 파일을 파싱하는 도중 에러가 발생했습니다: {e}")
                
    st.markdown("---")
    
    # --- 초기화 기능 레이아웃 ---
    st.markdown("### 🚨 기존 데이터 초기화 섹션")
    col_reset1, col_reset2 = st.columns(2)
    with col_reset1:
        st.markdown("#### 📅 당일 데이터 초기화")
        confirm_day = st.checkbox("정말로 오늘 데이터를 전부 삭제하는 것에 동의합니다.", key="confirm_day_reset")
        if st.button("🗑️ 당일 데이터 초기화 실행", type="primary", disabled=not confirm_day):
            if os.path.exists(STOCK_LOG_FILE):
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                shutil.copyfile(STOCK_LOG_FILE, f"vini_backup_{timestamp}.csv")
                logs = pd.read_csv(STOCK_LOG_FILE, encoding='utf-8-sig')
                filtered_logs = logs[logs["날짜"] != datetime.date.today().strftime("%Y-%m-%d")]
                filtered_logs.to_csv(STOCK_LOG_FILE, index=False, encoding='utf-8-sig')
                st.session_state.success_msg = "✅ 당일 로그 초기화 완료!"
                st.cache_data.clear()
                st.rerun()

    with col_reset2:
        st.markdown("#### 🗓️ 당월 데이터 초기화")
        confirm_month = st.checkbox("정말로 이번 달 데이터를 전부 삭제하는 것에 동의합니다.", key="confirm_month_reset")
        if st.button("💥 당월 데이터 전체 초기화 실행", type="primary", disabled=not confirm_month):
            if os.path.exists(STOCK_LOG_FILE):
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                shutil.copyfile(STOCK_LOG_FILE, f"vini_backup_{timestamp}.csv")
                logs = pd.read_csv(STOCK_LOG_FILE, encoding='utf-8-sig')
                logs['날짜_dt'] = pd.to_datetime(logs['날짜'])
                filtered_logs = logs[logs['날짜_dt'].dt.strftime("%Y-%m") != datetime.date.today().strftime("%Y-%m")]
                filtered_logs = filtered_logs.drop(columns=['날짜_dt'])
                filtered_logs.to_csv(STOCK_LOG_FILE, index=False, encoding='utf-8-sig')
                st.session_state.success_msg = "✅ 당월 로그 전체 초기화 완료!"
                st.cache_data.clear()
                st.rerun()
                
    st.markdown("---")
    st.markdown("### 🔥 서버 전체 공장 초기화")
    st.warning("과거에 생성되었던 데이터 백업 캐시를 완전히 강제 삭제하고, 시스템 내장 원본 엑셀 기준으로 정렬합니다.")
    
    confirm_destroy = st.checkbox("⚠️ 과거 기록된 로컬 데이터베이스를 전부 삭제하고 시스템 초기 파일 기준으로 정렬하는 것에 동의합니다.", key="confirm_destroy")
    if st.button("🚀 서버 강제 공장 초기화 실행", type="primary", disabled=not confirm_destroy):
        if os.path.exists(STOCK_LOG_FILE): os.remove(STOCK_LOG_FILE)  
        if os.path.exists(CUSTOM_MASTER_FILE): os.remove(CUSTOM_MASTER_FILE)  
        if os.path.exists(WINE_LOG_FILE): os.remove(WINE_LOG_FILE)
        if os.path.exists(WINE_MASTER_FILE): os.remove(WINE_MASTER_FILE)
        st.cache_data.clear()
        st.session_state.success_msg = f"💥 공장 초기화 완수! 기본 원본 엑셀 파일 데이터로 복구되었습니다."
        st.rerun()
